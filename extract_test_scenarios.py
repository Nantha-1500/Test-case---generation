from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook


REPO_ROOT = Path(__file__).resolve().parent
OUTPUTS_DIR = REPO_ROOT / "outputs"
TARGET_FILENAME = "MSME Base_TestScenarios_V1.xlsx"


def _latest_source_workbook(outputs_dir: Path) -> Path:
    candidates = [
        path
        for path in outputs_dir.glob("*.xlsx")
        if path.name != TARGET_FILENAME and "_TestScenarios_" not in path.name
    ]
    if not candidates:
        raise FileNotFoundError(f"No source workbook found under: {outputs_dir}")

    final_named = [path for path in candidates if "final" in path.name.lower()]
    pool = final_named or candidates
    return max(pool, key=lambda p: p.stat().st_mtime)


def _normalize_description(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return "Verify that"

    if text.startswith("Verify that"):
        suffix = text[len("Verify that") :].lstrip()
        return "Verify that" if not suffix else f"Verify that {suffix}"

    if text.lower().startswith("verify that"):
        suffix = text[len("verify that") :].lstrip()
        return "Verify that" if not suffix else f"Verify that {suffix}"

    return f"Verify that {text}"


def _iter_rows(source_path: Path) -> Iterable[tuple[object, object, object, object]]:
    workbook = load_workbook(source_path, data_only=True)
    for sheet in workbook.worksheets:
        headers = {}
        for col in range(1, sheet.max_column + 1):
            value = sheet.cell(1, col).value
            if isinstance(value, str):
                headers[value.strip().lower()] = col

        required = {"s.no", "screen name", "tc id", "test scenario description"}
        if not required.issubset(headers):
            continue

        for row in range(2, sheet.max_row + 1):
            s_no = sheet.cell(row, headers["s.no"]).value
            screen_name = sheet.cell(row, headers["screen name"]).value
            scenario_id = sheet.cell(row, headers["tc id"]).value
            description = sheet.cell(row, headers["test scenario description"]).value

            if not any([s_no, screen_name, scenario_id, description]):
                continue

            yield s_no, screen_name, scenario_id, _normalize_description(description)


def main() -> None:
    source_path = _latest_source_workbook(OUTPUTS_DIR)

    unique_rows: list[tuple[object, object, object, object]] = []
    seen: set[tuple[str, str, str]] = set()
    for s_no, screen_name, scenario_id, description in _iter_rows(source_path):
        key = (str(scenario_id or "").strip(), str(screen_name or "").strip(), description)
        if key in seen:
            continue
        seen.add(key)
        unique_rows.append((s_no, screen_name, scenario_id, description))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Test Scenarios"
    sheet.append(["S.No", "Screen Name", "Test Scenario", "Test Scenario Description"])
    for row in unique_rows:
        sheet.append(list(row))

    output_path = OUTPUTS_DIR / TARGET_FILENAME
    workbook.save(output_path)
    print(f"Source workbook: {source_path.name}")
    print(f"Rows written: {len(unique_rows)}")
    print(f"Output workbook: {output_path}")


if __name__ == "__main__":
    main()
