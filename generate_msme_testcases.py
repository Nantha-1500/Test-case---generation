#!/usr/bin/env python3
"""
MSME BASE Test Case Generator
Parses MSME BASE SRS V1.0.docx and 30 embedded Excel files to generate comprehensive test cases.
"""

import os
import sys
from docx import Document
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
from datetime import datetime

class MSMETestCaseGenerator:
    def __init__(self):
        self.docx_file = "requirements/MSME BASE SRS V1.0.docx"
        self.excel_files_dir = "temp_extraction/word/embeddings"
        self.template_file = "templates/testcase-template.xlsx"
        self.output_file = "outputs/MSME_BASE_TEST_CASES.xlsx"
        
        self.test_cases = []
        self.tc_counter = 1
        self.screens = []
        self.fields = {}
        self.validations = {}
        self.business_rules = []
        
    def parse_docx(self):
        """Parse MSME BASE SRS document for screens, workflows, and business rules."""
        print(f"Parsing {self.docx_file}...")
        try:
            doc = Document(self.docx_file)
            current_screen = None
            
            for para in doc.paragraphs:
                text = para.text.strip()
                if not text:
                    continue
                
                # Extract screens from headings
                if para.style.name.startswith('Heading'):
                    try:
                        level = int(para.style.name.replace('Heading ', '').replace('Heading', '').split('_')[0])
                        if level <= 3:
                            current_screen = text
                            if current_screen not in self.screens:
                                self.screens.append(current_screen)
                                print(f"  Found screen: {current_screen}")
                    except:
                        current_screen = text
                        if current_screen not in self.screens:
                            self.screens.append(current_screen)
                
                # Extract business rules
                if 'business rule' in text.lower() or 'validation' in text.lower():
                    self.business_rules.append({
                        'screen': current_screen,
                        'rule': text
                    })
                    
            print(f"Found {len(self.screens)} screens and {len(self.business_rules)} business rules")
            
        except Exception as e:
            print(f"Error parsing DOCX: {e}")
            sys.exit(1)
    
    def parse_excel_files(self):
        """Parse all 30 embedded Excel files for field details and validations."""
        print(f"\nParsing Excel files from {self.excel_files_dir}...")
        
        excel_files = sorted([f for f in os.listdir(self.excel_files_dir) if f.endswith('.xlsx')])
        print(f"Found {len(excel_files)} Excel files")
        
        for excel_file in excel_files:
            file_path = os.path.join(self.excel_files_dir, excel_file)
            print(f"  Processing {excel_file}...")
            
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    
                    # Try to extract field information
                    headers = []
                    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                        if row_idx == 1:
                            headers = [str(cell).strip() if cell else f"Col{idx}" for idx, cell in enumerate(row)]
                            continue
                        
                        if not any(row):
                            continue
                        
                        row_data = dict(zip(headers, row))
                        
                        # Extract field information
                        for key in ['Field Name', 'Field', 'Name', 'Control Name']:
                            if key in row_data and row_data[key]:
                                field_name = str(row_data[key]).strip()
                                if field_name and field_name != 'None':
                                    if field_name not in self.fields:
                                        self.fields[field_name] = {
                                            'source': excel_file,
                                            'sheet': sheet_name,
                                            'properties': row_data
                                        }
                                    break
                        
                        # Extract validations
                        for key in ['Validation', 'Validations', 'Rules', 'Mandatory']:
                            if key in row_data and row_data[key]:
                                validation = str(row_data[key]).strip()
                                if validation and validation != 'None':
                                    if sheet_name not in self.validations:
                                        self.validations[sheet_name] = []
                                    self.validations[sheet_name].append({
                                        'validation': validation,
                                        'data': row_data
                                    })
                
                wb.close()
                
            except Exception as e:
                print(f"    Error processing {excel_file}: {e}")
        
        print(f"Extracted {len(self.fields)} fields and {len(self.validations)} validation groups")
    
    def generate_test_id(self):
        """Generate test case ID in format MSME_BASE_TC001."""
        tc_id = f"MSME_BASE_TC{self.tc_counter:03d}"
        self.tc_counter += 1
        return tc_id
    
    def add_test_case(self, screen, scenario, description, procedure, expected):
        """Add a test case to the list."""
        self.test_cases.append({
            'S.No': len(self.test_cases) + 1,
            'Screen Name': screen,
            'Test Scenario': scenario,
            'Test Scenario Description': description,
            'Test Procedure': procedure,
            'Expected Output': expected,
            'Test Result': '',
            'QA Comments': ''
        })
    
    def generate_login_test_cases(self):
        """Generate test cases for Login functionality."""
        screen = "Login"
        
        # Positive tests
        self.add_test_case(
            screen, self.generate_test_id(),
            "Verify that user can login with valid credentials",
            "1. Navigate to login page\n2. Enter valid username\n3. Enter valid password\n4. Click Login button",
            "User is successfully logged in and redirected to Home Screen"
        )
        
        # Negative tests
        self.add_test_case(
            screen, self.generate_test_id(),
            "Verify that login fails with invalid username",
            "1. Navigate to login page\n2. Enter invalid username\n3. Enter valid password\n4. Click Login button",
            "Error message displayed: 'Invalid username or password'"
        )
        
        self.add_test_case(
            screen, self.generate_test_id(),
            "Verify that login fails with invalid password",
            "1. Navigate to login page\n2. Enter valid username\n3. Enter invalid password\n4. Click Login button",
            "Error message displayed: 'Invalid username or password'"
        )
        
        self.add_test_case(
            screen, self.generate_test_id(),
            "Verify that username field is mandatory",
            "1. Navigate to login page\n2. Leave username field empty\n3. Enter valid password\n4. Click Login button",
            "Error message displayed: 'Username is required'"
        )
        
        self.add_test_case(
            screen, self.generate_test_id(),
            "Verify that password field is mandatory",
            "1. Navigate to login page\n2. Enter valid username\n3. Leave password field empty\n4. Click Login button",
            "Error message displayed: 'Password is required'"
        )
    
    def generate_home_screen_test_cases(self):
        """Generate test cases for Home Screen."""
        screen = "Home Screen"
        
        self.add_test_case(
            screen, self.generate_test_id(),
            "Verify that Home Screen displays after successful login",
            "1. Login with valid credentials",
            "Home Screen is displayed with all menu options and widgets"
        )
        
        self.add_test_case(
            screen, self.generate_test_id(),
            "Verify that user can navigate to Inbox from Home Screen",
            "1. Login successfully\n2. Click on Inbox menu option",
            "Inbox screen is displayed"
        )
        
        self.add_test_case(
            screen, self.generate_test_id(),
            "Verify that user can logout from Home Screen",
            "1. Login successfully\n2. Click on Logout button",
            "User is logged out and redirected to Login page"
        )
    
    def generate_inbox_test_cases(self):
        """Generate test cases for Inbox functionality."""
        for inbox_type in ["Pending Approvals", "Rejected Applications", "Approved Applications"]:
            screen = f"Inbox >> {inbox_type}"
            
            self.add_test_case(
                screen, self.generate_test_id(),
                f"Verify that {inbox_type} inbox displays correctly",
                f"1. Login successfully\n2. Navigate to Inbox\n3. Select {inbox_type} tab",
                f"{inbox_type} list is displayed with all columns"
            )
            
            self.add_test_case(
                screen, self.generate_test_id(),
                f"Verify that search functionality works in {inbox_type}",
                f"1. Navigate to {inbox_type} inbox\n2. Enter search criteria\n3. Click Search button",
                "Filtered results are displayed based on search criteria"
            )
    
    def generate_create_application_test_cases(self):
        """Generate test cases for Create Application."""
        screen = "Create Application"
        
        self.add_test_case(
            screen, self.generate_test_id(),
            "Verify that user can create new MSME application",
            "1. Login successfully\n2. Click on Create Application\n3. Select MSME scheme\n4. Fill mandatory fields\n5. Click Submit",
            "Application is created successfully with unique application number"
        )
        
        self.add_test_case(
            screen, self.generate_test_id(),
            "Verify that application type selection is mandatory",
            "1. Navigate to Create Application\n2. Do not select application type\n3. Try to proceed",
            "Error message: 'Please select application type'"
        )
    
    def generate_party_details_test_cases(self):
        """Generate test cases for Party Details."""
        screen = "Party Details"
        
        # Generate test cases for common fields
        common_fields = [
            ("Party Name", "text", "Enter valid party name", "Party name is accepted"),
            ("PAN", "alphanumeric", "Enter valid PAN (AAAAA9999A)", "PAN is validated and accepted"),
            ("Mobile Number", "numeric", "Enter 10 digit mobile number", "Mobile number is accepted"),
            ("Email", "email", "Enter valid email address", "Email is validated and accepted")
        ]
        
        for field_name, field_type, action, expected in common_fields:
            # Positive test
            self.add_test_case(
                screen, self.generate_test_id(),
                f"Verify that {field_name} field accepts valid input",
                f"1. Navigate to Party Details\n2. {action}",
                expected
            )
            
            # Mandatory test
            self.add_test_case(
                screen, self.generate_test_id(),
                f"Verify that {field_name} field is mandatory",
                f"1. Navigate to Party Details\n2. Leave {field_name} field empty\n3. Try to proceed",
                f"Error message: '{field_name} is required'"
            )
    
    def generate_field_validation_test_cases(self):
        """Generate test cases for field validations from Excel files."""
        for field_name, field_info in self.fields.items():
            if len(self.test_cases) >= 500:
                break
                
            screen = "Field Validations"
            props = field_info['properties']
            
            # Mandatory validation
            if 'Mandatory' in props and props['Mandatory'] in ['Yes', 'Y', True, 'TRUE']:
                self.add_test_case(
                    screen, self.generate_test_id(),
                    f"Verify that {field_name} field is mandatory",
                    f"1. Navigate to screen containing {field_name}\n2. Leave field empty\n3. Try to proceed",
                    f"Error message displayed: '{field_name} is required'"
                )
            
            # Length validation
            if 'Max Length' in props and props['Max Length']:
                try:
                    max_len = int(props['Max Length'])
                    self.add_test_case(
                        screen, self.generate_test_id(),
                        f"Verify that {field_name} field accepts input within max length",
                        f"1. Navigate to screen\n2. Enter {max_len} characters in {field_name}",
                        f"{field_name} accepts input up to {max_len} characters"
                    )
                    
                    self.add_test_case(
                        screen, self.generate_test_id(),
                        f"Verify that {field_name} field rejects input exceeding max length",
                        f"1. Navigate to screen\n2. Enter {max_len + 1} characters in {field_name}",
                        f"Error message: 'Maximum {max_len} characters allowed'"
                    )
                except:
                    pass
    
    def generate_workflow_test_cases(self):
        """Generate workflow and integration test cases."""
        workflows = [
            ("Application Submission", "Create Application >> Submit", 
             "1. Create new application\n2. Fill all mandatory fields\n3. Upload required documents\n4. Click Submit",
             "Application submitted successfully and moved to Pending Approvals"),
            
            ("Application Approval", "Inbox >> Pending Approvals >> Approve",
             "1. Login as approver\n2. Navigate to Pending Approvals\n3. Select application\n4. Click Approve\n5. Add remarks\n6. Submit",
             "Application approved and moved to Approved Applications"),
            
            ("Application Rejection", "Inbox >> Pending Approvals >> Reject",
             "1. Login as approver\n2. Navigate to Pending Approvals\n3. Select application\n4. Click Reject\n5. Add rejection remarks\n6. Submit",
             "Application rejected and moved to Rejected Applications"),
        ]
        
        for name, screen, procedure, expected in workflows:
            self.add_test_case(
                screen, self.generate_test_id(),
                f"Verify that {name} workflow completes successfully",
                procedure,
                expected
            )
    
    def generate_all_test_cases(self):
        """Generate all test cases."""
        print("\nGenerating test cases...")
        
        # Generate test cases for different modules
        self.generate_login_test_cases()
        print(f"  Login: {len(self.test_cases)} test cases")
        
        self.generate_home_screen_test_cases()
        print(f"  Home Screen: {len(self.test_cases)} test cases")
        
        self.generate_inbox_test_cases()
        print(f"  Inbox: {len(self.test_cases)} test cases")
        
        self.generate_create_application_test_cases()
        print(f"  Create Application: {len(self.test_cases)} test cases")
        
        self.generate_party_details_test_cases()
        print(f"  Party Details: {len(self.test_cases)} test cases")
        
        self.generate_workflow_test_cases()
        print(f"  Workflows: {len(self.test_cases)} test cases")
        
        # Generate field validation test cases to reach 500+
        self.generate_field_validation_test_cases()
        print(f"  Field Validations: {len(self.test_cases)} test cases")
        
        # Generate additional test cases for each screen found
        for screen in self.screens[:50]:  # Limit to prevent too many
            if len(self.test_cases) >= 500:
                break
            
            self.add_test_case(
                screen, self.generate_test_id(),
                f"Verify that {screen} screen loads successfully",
                f"1. Login successfully\n2. Navigate to {screen}",
                f"{screen} screen is displayed with all fields and controls"
            )
            
            self.add_test_case(
                screen, self.generate_test_id(),
                f"Verify that mandatory fields in {screen} are validated",
                f"1. Navigate to {screen}\n2. Leave mandatory fields empty\n3. Try to proceed",
                "Appropriate error messages displayed for all mandatory fields"
            )
        
        print(f"\nTotal test cases generated: {len(self.test_cases)}")
    
    def create_excel_output(self):
        """Create Excel file with test cases and proper formatting."""
        print(f"\nCreating Excel output: {self.output_file}...")
        
        # Ensure outputs directory exists
        os.makedirs(os.path.dirname(self.output_file), exist_ok=True)
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create Test Cases sheet
        ws_tc = wb.create_sheet("Test Cases", 0)
        
        # Define headers
        headers = ['S.No', 'Screen Name', 'Test Scenario', 'Test Scenario Description', 
                   'Test Procedure', 'Expected Output', 'Test Result', 'QA Comments']
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws_tc.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Write test case data
        for row_idx, tc in enumerate(self.test_cases, 2):
            for col_idx, header in enumerate(headers, 1):
                value = tc.get(header, '')
                cell = ws_tc.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Auto-adjust column widths
        column_widths = {
            'A': 8,  # S.No
            'B': 30,  # Screen Name
            'C': 20,  # Test Scenario
            'D': 40,  # Test Scenario Description
            'E': 50,  # Test Procedure
            'F': 40,  # Expected Output
            'G': 15,  # Test Result
            'H': 30   # QA Comments
        }
        
        for col, width in column_widths.items():
            ws_tc.column_dimensions[col].width = width
        
        # Set row heights
        ws_tc.row_dimensions[1].height = 30
        for row in range(2, len(self.test_cases) + 2):
            ws_tc.row_dimensions[row].height = 60
        
        # Freeze panes
        ws_tc.freeze_panes = 'A2'
        
        # Add auto-filter
        ws_tc.auto_filter.ref = f"A1:H{len(self.test_cases) + 1}"
        
        # Create Coverage Summary sheet
        ws_summary = wb.create_sheet("Coverage Summary", 1)
        
        summary_headers = ['Module', 'Total Test Cases', 'Positive', 'Negative', 
                          'Mandatory', 'Validation', 'Workflow']
        
        for col_idx, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Calculate coverage statistics
        screen_stats = {}
        for tc in self.test_cases:
            screen = tc['Screen Name']
            if screen not in screen_stats:
                screen_stats[screen] = {'total': 0, 'positive': 0, 'negative': 0,
                                       'mandatory': 0, 'validation': 0, 'workflow': 0}
            screen_stats[screen]['total'] += 1
            
            desc = tc['Test Scenario Description'].lower()
            if 'valid' in desc and 'invalid' not in desc:
                screen_stats[screen]['positive'] += 1
            if 'invalid' in desc or 'fail' in desc or 'reject' in desc:
                screen_stats[screen]['negative'] += 1
            if 'mandatory' in desc:
                screen_stats[screen]['mandatory'] += 1
            if 'validation' in desc or 'format' in desc or 'length' in desc:
                screen_stats[screen]['validation'] += 1
            if 'workflow' in desc or '>>' in tc['Screen Name']:
                screen_stats[screen]['workflow'] += 1
        
        # Write summary data
        row_idx = 2
        for screen, stats in sorted(screen_stats.items()):
            ws_summary.cell(row=row_idx, column=1, value=screen)
            ws_summary.cell(row=row_idx, column=2, value=stats['total'])
            ws_summary.cell(row=row_idx, column=3, value=stats['positive'])
            ws_summary.cell(row=row_idx, column=4, value=stats['negative'])
            ws_summary.cell(row=row_idx, column=5, value=stats['mandatory'])
            ws_summary.cell(row=row_idx, column=6, value=stats['validation'])
            ws_summary.cell(row=row_idx, column=7, value=stats['workflow'])
            row_idx += 1
        
        # Add total row
        ws_summary.cell(row=row_idx, column=1, value="TOTAL").font = Font(bold=True)
        for col_idx in range(2, 8):
            total = sum(ws_summary.cell(row=r, column=col_idx).value or 0 
                       for r in range(2, row_idx))
            ws_summary.cell(row=row_idx, column=col_idx, value=total).font = Font(bold=True)
        
        # Format summary sheet
        for row in ws_summary.iter_rows(min_row=1, max_row=row_idx, min_col=1, max_col=7):
            for cell in row:
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust summary columns
        for col in range(1, 8):
            ws_summary.column_dimensions[get_column_letter(col)].width = 20
        
        # Save workbook
        wb.save(self.output_file)
        print(f"Excel file created successfully: {self.output_file}")
        print(f"  - Test Cases sheet: {len(self.test_cases)} test cases")
        print(f"  - Coverage Summary sheet: {len(screen_stats)} modules")
    
    def run(self):
        """Main execution method."""
        print("=" * 80)
        print("MSME BASE Test Case Generator")
        print("=" * 80)
        
        # Parse documents
        self.parse_docx()
        self.parse_excel_files()
        
        # Generate test cases
        self.generate_all_test_cases()
        
        # Create Excel output
        self.create_excel_output()
        
        print("\n" + "=" * 80)
        print("Test case generation completed successfully!")
        print("=" * 80)

if __name__ == "__main__":
    generator = MSMETestCaseGenerator()
    generator.run()
